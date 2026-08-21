"""
attendance_db.py
=================
Local SQLite storage, shared by both TA laptops. Each laptop keeps its own
attendance_local.db; the sync layer (sync_service.py) merges the two
together while class is in session.

Every function opens and closes its own short-lived connection (WAL mode +
a busy-timeout) instead of a connection being held open and passed between
threads. At this data scale (a few hundred students, a few thousand scans a
term) that overhead is negligible, and it avoids SQLite's "connection
created in one thread, used in another" restriction entirely -- both the
GUI thread and the background sync thread call these functions directly
and safely.

Tables:
    roster   -- uid -> student_id, name, faculty            (upsert, last-write-wins by updated_at)
    sessions -- session_id -> hall, ta_name, duration_minutes, closed_at  (upsert, last-write-wins)
    scans    -- immutable events (insert-only, deduped by id -- safe to re-merge any number of times)
"""

import sqlite3
from datetime import datetime

DB_FILE = "attendance_local.db"
EXPORT_FILE = "attendance_logger.xlsx"


def now_iso():
    return datetime.now().isoformat(timespec="microseconds")


def _conn(db_path):
    conn = sqlite3.connect(db_path, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA busy_timeout=10000;")
    return conn


def init_db(db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS roster (
                uid TEXT PRIMARY KEY,
                student_id TEXT,
                name TEXT,
                faculty TEXT,
                updated_at TEXT
            );
            CREATE TABLE IF NOT EXISTS sessions (
                session_id TEXT PRIMARY KEY,
                hall TEXT,
                ta_name TEXT,
                duration_minutes REAL,
                started_at TEXT,
                closed_at TEXT,
                updated_at TEXT
            );
            CREATE TABLE IF NOT EXISTS scans (
                id TEXT PRIMARY KEY,
                uid TEXT,
                student_id TEXT,
                name TEXT,
                faculty TEXT,
                session_id TEXT,
                source TEXT,
                ts TEXT
            );
            """
        )
        conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------- roster --
def upsert_roster(uid, student_id, name, faculty, db_path=DB_FILE, updated_at=None):
    updated_at = updated_at or now_iso()
    conn = _conn(db_path)
    try:
        conn.execute(
            "INSERT INTO roster (uid, student_id, name, faculty, updated_at) VALUES (?,?,?,?,?) "
            "ON CONFLICT(uid) DO UPDATE SET student_id=excluded.student_id, name=excluded.name, "
            "faculty=excluded.faculty, updated_at=excluded.updated_at "
            "WHERE excluded.updated_at >= roster.updated_at",
            (uid, student_id, name, faculty, updated_at),
        )
        conn.commit()
    finally:
        conn.close()


def get_roster(db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        return {row["uid"]: dict(row) for row in conn.execute("SELECT * FROM roster")}
    finally:
        conn.close()


def get_roster_entry(uid, db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        row = conn.execute("SELECT * FROM roster WHERE uid=?", (uid,)).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


# --------------------------------------------------------------- sessions --
def get_open_session(db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        row = conn.execute(
            "SELECT * FROM sessions WHERE duration_minutes IS NULL ORDER BY started_at DESC LIMIT 1"
        ).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def start_session(session_id, db_path=DB_FILE):
    now = now_iso()
    conn = _conn(db_path)
    try:
        conn.execute(
            "INSERT OR IGNORE INTO sessions (session_id, hall, ta_name, duration_minutes, started_at, closed_at, updated_at) "
            "VALUES (?, NULL, NULL, NULL, ?, NULL, ?)",
            (session_id, now, now),
        )
        conn.commit()
    finally:
        conn.close()


def close_session(session_id, hall, ta_name, duration_minutes, db_path=DB_FILE):
    now = now_iso()
    conn = _conn(db_path)
    try:
        conn.execute(
            "UPDATE sessions SET hall=?, ta_name=?, duration_minutes=?, closed_at=?, updated_at=? WHERE session_id=?",
            (hall, ta_name, duration_minutes, now, now, session_id),
        )
        conn.commit()
    finally:
        conn.close()


def get_all_sessions(db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        return {row["session_id"]: dict(row) for row in conn.execute("SELECT * FROM sessions")}
    finally:
        conn.close()


def highest_session_number(date_nodash, db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        rows = conn.execute(
            "SELECT session_id FROM sessions WHERE session_id LIKE ?", (f"{date_nodash}_%",)
        ).fetchall()
    finally:
        conn.close()
    best = 0
    for row in rows:
        try:
            best = max(best, int(row["session_id"].rsplit("_", 1)[1]))
        except (ValueError, IndexError):
            pass
    return best


# ------------------------------------------------------------------ scans --
def insert_scan(scan_id, uid, student_id, name, faculty, session_id, source, db_path=DB_FILE, ts=None):
    ts = ts or now_iso()
    conn = _conn(db_path)
    try:
        conn.execute(
            "INSERT OR IGNORE INTO scans (id, uid, student_id, name, faculty, session_id, source, ts) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (scan_id, uid, student_id, name, faculty, session_id, source, ts),
        )
        conn.commit()
        return conn.total_changes
    finally:
        conn.close()


def get_all_scans(db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        return [dict(row) for row in conn.execute("SELECT * FROM scans")]
    finally:
        conn.close()


def get_recent_scans(limit=50, db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        rows = conn.execute("SELECT * FROM scans ORDER BY ts DESC LIMIT ?", (limit,)).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ------------------------------------------------------------- sync merge --
def export_all(db_path=DB_FILE):
    conn = _conn(db_path)
    try:
        return {
            "roster": [dict(r) for r in conn.execute("SELECT * FROM roster")],
            "sessions": [dict(r) for r in conn.execute("SELECT * FROM sessions")],
            "scans": [dict(r) for r in conn.execute("SELECT * FROM scans")],
        }
    finally:
        conn.close()


def merge_payload(payload, db_path=DB_FILE):
    """Merge a peer's full export into this local database. Returns the
    number of NEW scan rows that weren't already present locally."""
    conn = _conn(db_path)
    added_scans = 0
    try:
        for row in payload.get("scans", []):
            before = conn.execute("SELECT 1 FROM scans WHERE id=?", (row["id"],)).fetchone()
            conn.execute(
                "INSERT OR IGNORE INTO scans (id, uid, student_id, name, faculty, session_id, source, ts) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (
                    row["id"], row["uid"], row["student_id"], row["name"],
                    row["faculty"], row["session_id"], row["source"], row["ts"],
                ),
            )
            if before is None:
                added_scans += 1

        for row in payload.get("sessions", []):
            conn.execute(
                "INSERT INTO sessions (session_id, hall, ta_name, duration_minutes, started_at, closed_at, updated_at) "
                "VALUES (?,?,?,?,?,?,?) "
                "ON CONFLICT(session_id) DO UPDATE SET hall=excluded.hall, ta_name=excluded.ta_name, "
                "duration_minutes=excluded.duration_minutes, started_at=excluded.started_at, "
                "closed_at=excluded.closed_at, updated_at=excluded.updated_at "
                "WHERE excluded.updated_at >= sessions.updated_at",
                (
                    row["session_id"], row["hall"], row["ta_name"], row["duration_minutes"],
                    row["started_at"], row["closed_at"], row["updated_at"],
                ),
            )

        for row in payload.get("roster", []):
            conn.execute(
                "INSERT INTO roster (uid, student_id, name, faculty, updated_at) VALUES (?,?,?,?,?) "
                "ON CONFLICT(uid) DO UPDATE SET student_id=excluded.student_id, name=excluded.name, "
                "faculty=excluded.faculty, updated_at=excluded.updated_at "
                "WHERE excluded.updated_at >= roster.updated_at",
                (row["uid"], row["student_id"], row["name"], row["faculty"], row["updated_at"]),
            )

        conn.commit()
    finally:
        conn.close()
    return added_scans


# ---------------------------------------------------------------- export --
def export_to_xlsx(db_path=DB_FILE, path=EXPORT_FILE):
    """Flatten scans+sessions into the row format attendance_analyzer.py
    already expects: one sheet per date, columns
    Time | UID | Student ID | Name | Faculty | Session ID | Hall Number | TA Name | Duration (min)"""
    import openpyxl

    sessions = get_all_sessions(db_path)
    scans = get_all_scans(db_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = {}

    for scan in sorted(scans, key=lambda s: s["ts"]):
        ts = datetime.fromisoformat(scan["ts"])
        date_str = ts.strftime("%Y-%m-%d")
        if date_str not in sheets:
            ws = wb.create_sheet(date_str)
            ws.append(["Time", "UID", "Student ID", "Name", "Faculty", "Session ID", "Hall Number", "TA Name", "Duration (min)"])
            widths = {"A": 12, "B": 14, "C": 14, "D": 30, "E": 26, "F": 16, "G": 14, "H": 20, "I": 16}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            sheets[date_str] = ws
        ws = sheets[date_str]
        sess = sessions.get(scan["session_id"], {})
        ws.append(
            [
                ts.strftime("%H:%M:%S"),
                scan["uid"],
                scan["student_id"],
                scan["name"],
                scan["faculty"],
                scan["session_id"],
                sess.get("hall"),
                sess.get("ta_name"),
                sess.get("duration_minutes"),
            ]
        )

    if not sheets:
        wb.create_sheet("Sheet1")

    wb.save(path)
    return path