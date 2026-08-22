"""
Attendance Analyzer  (Phase II - run ONCE, after all halls' logs are merged)
=============================================================================
Each hall runs attendance_recorder.py locally and produces its own
attendance_logger.xlsx. At day's end, every hall's file is merged into one
(e.g. uploaded into a shared Google Sheet, then downloaded back as a single
attendance_logger.xlsx). This script is then run ONCE, by the lecturer or a
senior TA, on that merged file to produce the attendance report -- one sheet
per date found in the file, in chronological order.

INPUT FORMAT: every non-"Roster" sheet is read and EVERY scan row is placed
onto its date by parsing the date out of its own Session ID column
(YYYYMMDD_N) -- NOT by which sheet or tab the row happens to sit on. This is
deliberate: attendance_recorder.py's normal export writes one sheet per
date (sheet named "YYYY-MM-DD"), but this script doesn't actually require
that layout -- a single flattened sheet containing every day's rows works
identically, since the Session ID column is the one thing that's always
present and unambiguous per scan. That matters in practice because a
cross-hall merge (copy/paste into one shared Google Sheet, a manual
export, a hand-built test file) can easily lose the "one tab per date"
structure without losing any data -- this script tolerates that.

Merged file columns (per scan row):
    Time | UID | Student ID | Name | Faculty | Session ID | Hall Number | TA Name | Duration (min)

Status rules, evaluated per (student, session_id):

    - The student's scans for that session_id span MORE THAN ONE Hall Number
        -> "Needs Review - Hall Conflict"  (time spent left blank)
    - Otherwise, an odd number of scans for that student+session
        -> "Needs Review - Scan Missing"  (time spent left blank)
    - Otherwise, pair scans as (in, out), (in, out)... and sum the time
        -> "Attended"  if total minutes >= session duration x ATTENDANCE_RATIO
        -> "Absent"    otherwise
      Either way, the actual minutes spent in the hall is recorded.

IMPORTANT: duration is resolved per (session_id, HALL) pair, not per
session_id alone. The same session number can legitimately run with a
different actual duration in each hall (e.g. Session 1 takes 60 minutes in
Hall 1 but only 45 in Hall 2, because the two instructors wrapped up at
different times) -- each student is evaluated against the duration
recorded for the specific hall THEY were in, not some single figure
blended across every hall running that session number.

A student's roster info (for labeling, and for catching students who
never scanned at all) comes from the "Roster" sheet inside the SAME merged
attendance_logger.xlsx -- there's no separate roster.json to keep in sync;
attendance_recorder.py's export writes that sheet automatically.

Requirements:
    pip install openpyxl

Usage:
    python attendance_analyzer.py
        -> reports every date found in the file, chronologically, one sheet each
    python attendance_analyzer.py --input attendance_logger.xlsx --date 2026-08-19
        -> reports only that one date
"""

import argparse
import os
from collections import Counter
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

ATTENDANCE_FILE_DEFAULT = "attendance_logger.xlsx"
REPORT_PREFIX = "attendance_report_"

# A student must be inside the hall for at least this fraction of the
# session's actual (entered) duration, evaluated per hall (see module docstring).
ATTENDANCE_RATIO = 0.75


def load_roster(recorder_path):
    """Read the 'Roster' sheet (Card UID | Student ID | Name | Faculty)
    from the merged attendance_logger.xlsx. Returns {} if the file or that
    sheet doesn't exist -- older exports without the sheet still work, they
    just can't show fully-absent (zero-scan) students."""
    roster = {}
    if not os.path.exists(recorder_path):
        return roster
    wb = openpyxl.load_workbook(recorder_path, data_only=True)
    if "Roster" not in wb.sheetnames:
        return roster
    ws = wb["Roster"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        uid, student_id, name, faculty = (row + (None,) * 4)[:4]
        uid = str(uid).strip()
        if not uid:
            continue
        roster[uid] = {"student_id": student_id or uid, "name": name or "", "faculty": faculty or ""}
    return roster


def session_number(session_id):
    """'20260819_2' -> 2"""
    return int(session_id.rsplit("_", 1)[1])


def session_date_dash(session_id):
    """'20260819_2' -> '2026-08-19'. Raises ValueError on a malformed id."""
    raw = session_id.rsplit("_", 1)[0]
    return datetime.strptime(raw, "%Y%m%d").strftime("%Y-%m-%d")


def _parse_time_cell(time_val, sheet_date):
    if isinstance(time_val, datetime):
        return time_val
    if hasattr(time_val, "hour"):  # datetime.time
        return datetime.combine(sheet_date, time_val)
    return datetime.combine(sheet_date, datetime.strptime(str(time_val), "%H:%M:%S").time())


def read_all_scans(recorder_path):
    """Read every sheet except 'Roster' and bucket each scan row onto the
    date parsed from ITS OWN Session ID -- regardless of which sheet it's
    physically sitting on. Works equally well whether the file has one
    sheet per date (the normal export) or every day flattened into one
    sheet.

    Returns {date_str: (sessions, hall_durations)} where, per date:
      sessions: {session_id: {uid: {student_id, name, faculty, times, halls}}}
      hall_durations: {session_id: {hall: [duration values recorded on rows
                        with that session_id + hall]}}
    """
    by_date = {}
    if not os.path.exists(recorder_path):
        return by_date

    wb = openpyxl.load_workbook(recorder_path, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name == "Roster":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            padded = (row + (None,) * 9)[:9]
            time_val, uid, student_id, name, faculty, session_id, hall, ta_name, duration = padded
            if uid is None or not session_id:
                continue
            uid = str(uid).strip()
            session_id = str(session_id).strip()

            try:
                date_str = session_date_dash(session_id)
            except (ValueError, IndexError):
                continue  # Session ID doesn't encode a parseable date -- can't place this row anywhere, skip it

            sessions, hall_durations = by_date.setdefault(date_str, ({}, {}))
            sheet_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            ts = _parse_time_cell(time_val, sheet_date)

            sess = sessions.setdefault(session_id, {})
            entry = sess.setdefault(
                uid, {"student_id": student_id or "", "name": name or "", "faculty": faculty or "", "times": [], "halls": set()}
            )
            if student_id:
                entry["student_id"] = student_id
            if name and name != "UNKNOWN CARD":
                entry["name"] = name
            if faculty:
                entry["faculty"] = faculty
            entry["times"].append(ts)
            if hall:
                entry["halls"].add(hall)
                if duration not in (None, ""):
                    try:
                        hall_durations.setdefault(session_id, {}).setdefault(hall, []).append(float(duration))
                    except (TypeError, ValueError):
                        pass

    return by_date


def list_available_dates(recorder_path):
    """Every date any scan's Session ID resolves to, sorted chronologically."""
    return sorted(read_all_scans(recorder_path).keys())


def resolve_hall_duration(session_id, hall, hall_durations):
    """Most common Duration value entered for this specific session+hall
    combination. Returns None if nothing was ever recorded for it."""
    values = hall_durations.get(session_id, {}).get(hall, [])
    if not values:
        return None
    return Counter(values).most_common(1)[0][0]


def evaluate_student(info, threshold_minutes):
    """Returns (status, minutes_spent). minutes_spent is None whenever the
    status is a manual-review case -- there's nothing trustworthy to report."""
    if len(info["halls"]) > 1:
        return "Needs Review - Hall Conflict", None

    times = sorted(info["times"])
    n = len(times)
    if n % 2 != 0:
        return "Needs Review - Scan Missing", None

    total_seconds = 0.0
    for i in range(0, n, 2):
        total_seconds += (times[i + 1] - times[i]).total_seconds()
    total_minutes = round(total_seconds / 60.0, 1)
    status = "Attended" if total_minutes >= threshold_minutes else "Absent"
    return status, total_minutes


def build_report_rows(sessions, hall_durations, date_str, roster):
    """Returns (rows, ordered_ids, warnings) for one date's already-parsed
    sessions/hall_durations (see read_all_scans)."""
    ordered_ids = sorted(sessions.keys(), key=session_number)

    warnings = []
    for sid in ordered_ids:
        for hall, values in hall_durations.get(sid, {}).items():
            distinct = sorted(set(values))
            if len(distinct) > 1:
                chosen = Counter(values).most_common(1)[0][0]
                warnings.append(
                    f"{date_str} session {sid}, {hall}: inconsistent durations entered {distinct}, using {chosen:g} min."
                )

    all_uids = set(roster.keys())
    for sid in ordered_ids:
        all_uids |= set(sessions[sid].keys())

    def label(uid):
        if uid in roster:
            return roster[uid].get("student_id", uid), roster[uid].get("name", "UNKNOWN CARD")
        for sid in ordered_ids:
            info = sessions[sid].get(uid)
            if info and info["name"]:
                return info["student_id"], info["name"]
        return uid, "UNKNOWN CARD"

    rows = []
    for uid in sorted(all_uids):
        student_id, name = label(uid)
        row = {"uid": uid, "student_id": student_id, "name": name, "cells": {}}
        for sid in ordered_ids:
            info = sessions[sid].get(uid, {"times": [], "halls": set()})
            halls = info["halls"]
            if len(halls) > 1:
                status, minutes = "Needs Review - Hall Conflict", None
            else:
                hall = next(iter(halls), None)
                dur = resolve_hall_duration(sid, hall, hall_durations) if hall else None
                if dur is None:
                    if not info["times"]:
                        status, minutes = "Absent", 0.0  # enrolled, zero scans -- an unambiguous absence
                    else:
                        status, minutes = "Needs Review - Duration Missing", None  # scans exist but no duration was ever recorded for this hall
                else:
                    status, minutes = evaluate_student(info, dur * ATTENDANCE_RATIO)
            row["cells"][sid] = (minutes, status)
        rows.append(row)

    rows.sort(key=lambda r: (r["name"] == "UNKNOWN CARD", r["student_id"]))
    return rows, ordered_ids, warnings


def _duration_header_lines(ordered_ids, hall_durations):
    lines = []
    for idx, sid in enumerate(ordered_ids, start=1):
        per_hall = hall_durations.get(sid, {})
        if not per_hall:
            lines.append(f"Session {idx} Duration: unknown (no Duration recorded)")
            continue
        parts = [f"{hall} = {Counter(vals).most_common(1)[0][0]:g} min" for hall, vals in sorted(per_hall.items())]
        lines.append(f"Session {idx} Duration: " + ", ".join(parts))
    return lines


def _write_date_sheet(wb, date_str, rows, ordered_ids, hall_durations):
    ws = wb.create_sheet(date_str)
    ws.append([f"Attendance Report - {date_str}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    for line in _duration_header_lines(ordered_ids, hall_durations):
        ws.append([line])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
    ws.append([])

    header_row1 = ["Card UID", "Student ID", "Name"]
    header_row2 = ["", "", ""]
    for sid in ordered_ids:
        header_row1 += [sid, ""]
        header_row2 += ["time spent in hall", "status"]
    ws.append(header_row1)
    ws.append(header_row2)

    header_row1_idx = ws.max_row - 1
    col = 4
    for _ in ordered_ids:
        ws.merge_cells(start_row=header_row1_idx, start_column=col, end_row=header_row1_idx, end_column=col + 1)
        col += 2

    for row in rows:
        line = [row["uid"], row["student_id"], row["name"]]
        for sid in ordered_ids:
            minutes, status = row["cells"][sid]
            line += [minutes, status]
        ws.append(line)

    widths = {"A": 16, "B": 14, "C": 26}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    col = 4
    for _ in ordered_ids:
        ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions[get_column_letter(col + 1)].width = 22
        col += 2


def default_report_filename(dates):
    """Same naming convention used by both the CLI and the GUI, so there's
    one place that decides it: 'attendance_report_<date>.xlsx' for a single
    date, or '..._<first>_to_<last>.xlsx' for a range."""
    dates = sorted(dates)
    if len(dates) == 1:
        return f"{REPORT_PREFIX}{dates[0]}.xlsx"
    return f"{REPORT_PREFIX}{dates[0]}_to_{dates[-1]}.xlsx"


def generate_full_report(recorder_path, roster, dates=None, report_path=None):
    """One workbook, one sheet per date, sheets in chronological order.
    Pass dates=[single_date] to report just that one date."""
    all_data = read_all_scans(recorder_path)
    if not all_data:
        raise ValueError(
            "No scan data with a recognizable Session ID (format YYYYMMDD_N) was found in the input file."
        )
    available_dates = sorted(all_data.keys())

    if dates:
        missing = [d for d in dates if d not in available_dates]
        if missing:
            raise ValueError(
                f"Date(s) not found in the input file: {', '.join(missing)}. "
                f"Dates present: {', '.join(available_dates)}"
            )
        dates = sorted(dates)
    else:
        dates = available_dates

    if report_path is None:
        report_path = default_report_filename(dates)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    per_date_rows = {}
    all_warnings = []
    counts = {}
    for date_str in dates:
        sessions, hall_durations = all_data[date_str]
        rows, ordered_ids, warnings = build_report_rows(sessions, hall_durations, date_str, roster)
        _write_date_sheet(wb, date_str, rows, ordered_ids, hall_durations)
        per_date_rows[date_str] = (rows, ordered_ids)
        all_warnings.extend(warnings)
        for row in rows:
            for _minutes, status in row["cells"].values():
                counts[status] = counts.get(status, 0) + 1

    wb.save(report_path)
    return report_path, per_date_rows, all_warnings, counts


def main():
    p = argparse.ArgumentParser(description="Generate the attendance report from the merged log.")
    p.add_argument("--input", default=ATTENDANCE_FILE_DEFAULT, help="Path to the (merged) attendance_logger.xlsx")
    p.add_argument(
        "--date", default=None, help="Report a single date, YYYY-MM-DD. Omit to report every date found, chronologically."
    )
    p.add_argument("--output", default=None, help="Output report path")
    args = p.parse_args()

    if not os.path.exists(args.input):
        raise SystemExit(f"Input file not found: {args.input}")

    roster = load_roster(args.input)
    dates = [args.date] if args.date else None
    report_path, per_date_rows, warnings, counts = generate_full_report(args.input, roster, dates=dates, report_path=args.output)

    print(f"Report saved to: {report_path}")
    print("Dates covered:", ", ".join(per_date_rows.keys()))
    print("Totals:", ", ".join(f"{k}: {v}" for k, v in counts.items()) if counts else "no data found")
    for w in warnings:
        print(f"WARNING: {w}")


if __name__ == "__main__":
    main()