"""
Card Attendance Recorder  (Phase I - runs on each hall's PC during class)
============================================================================
Works with the OMNIKEY 5427 G2 in keyboard-emulation (HID) mode: the reader
"types" an 18-character hex string + Enter into whatever window has focus.
The first 8 characters change on every read; the last 10 are the card's
fixed UID, so only the last 10 characters are kept.

Two scanners per hall need NO special handling here: both are just
keyboard-emulation devices sending keystrokes to whichever window/field is
focused on this PC, so a student tapping in on one scanner and out on
another still increments the same in/out count for that card.

Enroll Cards tab now looks students up in a university-provided roster
(students_database.xlsx, columns: ID | Student Name | Faculty | ...) instead
of requiring every one of ~400 students to be typed in by hand:
    - Scan the card, type the printed Student ID, click Search (or press
      Enter) -- if that ID is in the database, Name and Faculty are filled
      in automatically.
    - If the ID isn't found, Name is typed manually and Faculty is chosen
      from the dropdown (both fields stay editable either way, so a
      database hit can still be corrected before saving).

This version writes a purely LOCAL log per hall/session -- it no longer
generates the attendance report itself. Once class ends, this hall's
attendance_logger.xlsx gets uploaded/merged with every other hall's file
(e.g. via a shared Google Sheet), and attendance_analyzer.py is run ONCE
on the merged file by the lecturer/senior TA (see that script's docstring).

Requirements:
    pip install openpyxl

Usage:
    python attendance_recorder.py

Files used/created next to this script:
    students_database.xlsx -- university-provided roster (read-only input)
    roster.json             -- uid -> {student_id, name, faculty}   (this course's enrolled cards)
    session_state.json      -- today's session counter + any active session
    attendance_logger.xlsx  -- raw scan log, one sheet per date, columns:
        Time | UID | Student ID | Name | Faculty | Session ID | Hall Number | TA Name | Duration (min)
"""

import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
from datetime import datetime
import openpyxl

ROSTER_FILE = "roster.json"
ATTENDANCE_FILE = "attendance_logger.xlsx"
SESSION_STATE_FILE = "session_state.json"
STUDENT_DB_FILE = "students_database.xlsx"
DEBOUNCE_SECONDS = 3
UID_LENGTH = 10
TYPICAL_SESSION_RANGE = (45, 90)  # minutes; soft sanity check only

# TODO: replace with the official list of halls once provided.
HALL_LIST = ["Hall 1", "Hall 2", "Hall 3"]

FACULTY_LIST = [
    "Biotechnology",
    "Business Administration",
    "Business Informatics",
    "Engineering",
    "Informatics and Computer Science",
    "Pharmaceutical Engineering",
]

# Column layout used in attendance_logger.xlsx (1-based, for openpyxl cells)
COL_TIME, COL_UID, COL_STUDENT_ID, COL_NAME, COL_FACULTY = 1, 2, 3, 4, 5
COL_SESSION_ID, COL_HALL, COL_TA, COL_DURATION = 6, 7, 8, 9


def extract_uid(raw_scan):
    raw_scan = raw_scan.strip().upper()
    if len(raw_scan) >= UID_LENGTH:
        return raw_scan[-UID_LENGTH:]
    return raw_scan


def load_roster():
    if os.path.exists(ROSTER_FILE):
        with open(ROSTER_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_roster(roster):
    with open(ROSTER_FILE, "w", encoding="utf-8") as f:
        json.dump(roster, f, indent=2, ensure_ascii=False)


def load_student_database(path=STUDENT_DB_FILE):
    """Read the university-provided roster: ID | Student Name | Faculty | ...
    Extra columns are ignored. Returns {student_id: {"name":.., "faculty":..}}."""
    db = {}
    if not os.path.exists(path):
        return db
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        padded = (row + (None,) * 3)[:3]
        sid, name, faculty = padded
        sid = str(sid).strip()
        if not sid:
            continue
        db[sid] = {
            "name": str(name).strip() if name is not None else "",
            "faculty": str(faculty).strip() if faculty is not None else "",
        }
    return db


def load_session_state():
    if os.path.exists(SESSION_STATE_FILE):
        with open(SESSION_STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"date": None, "next_session_number": 1, "active_session": None}


def save_session_state(state):
    with open(SESSION_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, ensure_ascii=False)


def ask_end_session_params(parent, title, initial_instructor="", initial_duration="", initial_hall=None):
    """Modal dialog for (instructor name, duration in minutes, hall number).
    Returns (instructor, duration, hall) or None if cancelled."""
    dialog = tk.Toplevel(parent)
    dialog.title(title)
    dialog.transient(parent)
    dialog.grab_set()
    result = {}

    ttk.Label(dialog, text="Instructor / TA name:").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 2))
    instructor_var = tk.StringVar(value=initial_instructor)
    instructor_entry = ttk.Entry(dialog, textvariable=instructor_var, width=32)
    instructor_entry.grid(row=1, column=0, padx=8, pady=(0, 8))

    ttk.Label(dialog, text="Session duration (minutes):").grid(row=2, column=0, sticky="w", padx=8, pady=(0, 2))
    duration_var = tk.StringVar(value=str(initial_duration) if initial_duration != "" else "")
    duration_entry = ttk.Entry(dialog, textvariable=duration_var, width=32)
    duration_entry.grid(row=3, column=0, padx=8, pady=(0, 8))

    ttk.Label(dialog, text="Hall number:").grid(row=4, column=0, sticky="w", padx=8, pady=(0, 2))
    hall_var = tk.StringVar(value=initial_hall or (HALL_LIST[0] if HALL_LIST else ""))
    hall_combo = ttk.Combobox(dialog, textvariable=hall_var, values=HALL_LIST, state="readonly", width=30)
    hall_combo.grid(row=5, column=0, padx=8, pady=(0, 8))

    def on_ok():
        name = instructor_var.get().strip()
        dur_raw = duration_var.get().strip()
        hall = hall_var.get().strip()
        if not name:
            messagebox.showwarning("Missing name", "Enter the instructor/TA name.", parent=dialog)
            return
        if not hall:
            messagebox.showwarning("Missing hall", "Select the hall number.", parent=dialog)
            return
        try:
            duration = float(dur_raw)
            if duration <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning(
                "Invalid duration", "Enter the session duration in minutes as a positive number.", parent=dialog
            )
            return
        lo, hi = TYPICAL_SESSION_RANGE
        if not (lo <= duration <= hi):
            if not messagebox.askyesno(
                "Unusual duration",
                f"{duration:g} minutes is outside the typical {lo}-{hi} minute range. Continue anyway?",
                parent=dialog,
            ):
                return
        result["instructor"] = name
        result["duration"] = duration
        result["hall"] = hall
        dialog.destroy()

    btn_row = ttk.Frame(dialog)
    btn_row.grid(row=6, column=0, pady=(0, 8))
    ttk.Button(btn_row, text="OK", command=on_ok).pack(side="left", padx=4)
    ttk.Button(btn_row, text="Cancel", command=dialog.destroy).pack(side="left", padx=4)

    instructor_entry.focus_set()
    dialog.wait_window()
    if "duration" not in result:
        return None
    return result["instructor"], result["duration"], result["hall"]


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Card Attendance Recorder")
        self.geometry("780x600")

        self.roster = load_roster()
        self.student_db = load_student_database()
        self.last_scan = {}
        self.last_instructor = ""
        self.last_hall = None

        self.session_state = load_session_state()
        today_nodash = datetime.now().strftime("%Y%m%d")
        if self.session_state.get("date") != today_nodash:
            if self.session_state.get("active_session"):
                messagebox.showwarning(
                    "Unclosed session from a previous day",
                    f"Session {self.session_state['active_session']['session_id']} was never closed. "
                    "It's being discarded for today -- review its scans manually if needed.",
                )
            self.session_state = {"date": today_nodash, "next_session_number": 1, "active_session": None}
            save_session_state(self.session_state)
        self.active_session = self.session_state.get("active_session")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self.attendance_tab = ttk.Frame(self.notebook)
        self.enroll_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.attendance_tab, text="Attendance")
        self.notebook.add(self.enroll_tab, text="Enroll Cards")

        self._build_attendance_tab()
        self._build_enroll_tab()
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self._update_session_ui()

        if not self.student_db:
            messagebox.showinfo(
                "Student database not found",
                f"'{STUDENT_DB_FILE}' wasn't found next to this script.\n"
                "Enroll Cards will still work, just with manual entry only "
                "(no auto-fill from the university roster).",
            )

    # ---------------- Attendance tab ----------------
    def _build_attendance_tab(self):
        frame = self.attendance_tab

        session_row = ttk.Frame(frame)
        session_row.pack(fill="x", padx=8, pady=8)
        self.start_btn = ttk.Button(session_row, text="Start Session", command=self._start_session)
        self.start_btn.pack(side="left")
        self.end_btn = ttk.Button(session_row, text="End Session", command=self._end_session)
        self.end_btn.pack(side="left", padx=(6, 0))
        self.session_status_label = ttk.Label(session_row, text="No active session", font=("Segoe UI", 10, "bold"))
        self.session_status_label.pack(side="left", padx=12)

        ttk.Label(frame, text="Scan a card (keep this tab open and this box focused):").pack(
            anchor="w", padx=8, pady=(4, 2)
        )
        self.att_entry = ttk.Entry(frame, font=("Consolas", 14))
        self.att_entry.pack(fill="x", padx=8)
        self.att_entry.bind("<Return>", self._on_attendance_scan)

        self.att_log = tk.Listbox(frame, font=("Consolas", 11))
        self.att_log.pack(fill="both", expand=True, padx=8, pady=8)

        self.att_status = ttk.Label(frame, text=f"Logging to: {ATTENDANCE_FILE}")
        self.att_status.pack(anchor="w", padx=8, pady=(0, 8))

    def _on_attendance_scan(self, event):
        raw = self.att_entry.get()
        self.att_entry.delete(0, tk.END)
        uid = extract_uid(raw)
        if not uid:
            return

        if uid not in self.roster:
            messagebox.showwarning(
                "Unregistered card",
                "This card is not registered in the system.\n"
                "Please add the student in the 'Enroll Cards' tab first.",
            )
            return
        if not self.active_session:
            messagebox.showwarning("No active session", "Click 'Start Session' before scanning cards.")
            return

        self._register_attendance(uid)

    def _register_attendance(self, uid):
        """Shared by live scans and enroll-time logging. No-op if no session is running."""
        if not self.active_session:
            return
        now = datetime.now()
        last = self.last_scan.get(uid)
        if last and (now - last).total_seconds() < DEBOUNCE_SECONDS:
            return
        self.last_scan[uid] = now

        entry = self.roster.get(uid, {})
        student_id = entry.get("student_id", "")
        name = entry.get("name", "UNKNOWN CARD")
        faculty = entry.get("faculty", "")
        session_id = self.active_session["session_id"]

        line = f"{now.strftime('%H:%M:%S')}   {uid}   {student_id:<10}   {name:<20}   [{session_id}]"
        self.att_log.insert(0, line)

        try:
            self._append_attendance_row(now, uid, student_id, name, faculty, session_id)
        except PermissionError:
            messagebox.showerror(
                "File is open",
                f"Couldn't save to {ATTENDANCE_FILE} because it's open in Excel.\n"
                "Close it and the next scan will save normally.",
            )

    def _append_attendance_row(self, ts, uid, student_id, name, faculty, session_id):
        if os.path.exists(ATTENDANCE_FILE):
            wb = openpyxl.load_workbook(ATTENDANCE_FILE)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        sheet_name = ts.strftime("%Y-%m-%d")
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(
                ["Time", "UID", "Student ID", "Name", "Faculty", "Session ID", "Hall Number", "TA Name", "Duration (min)"]
            )
            widths = {"A": 12, "B": 14, "C": 14, "D": 30, "E": 26, "F": 16, "G": 14, "H": 20, "I": 16}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
        else:
            ws = wb[sheet_name]

        # Hall Number / TA Name / Duration are filled in later, at "End Session".
        ws.append([ts.strftime("%H:%M:%S"), uid, student_id, name, faculty, session_id, None, None, None])
        wb.save(ATTENDANCE_FILE)

    def _apply_session_metadata(self, session_id, hall, instructor, duration):
        """Retroactively tag every already-written row of this session with
        the hall/instructor/duration collected at End Session time."""
        if not os.path.exists(ATTENDANCE_FILE):
            return 0
        wb = openpyxl.load_workbook(ATTENDANCE_FILE)
        date_str = datetime.strptime(session_id.split("_")[0], "%Y%m%d").strftime("%Y-%m-%d")
        if date_str not in wb.sheetnames:
            return 0
        ws = wb[date_str]

        updated = 0
        for row in ws.iter_rows(min_row=2):
            if row[COL_SESSION_ID - 1].value == session_id:
                row[COL_HALL - 1].value = hall
                row[COL_TA - 1].value = instructor
                row[COL_DURATION - 1].value = duration
                updated += 1

        wb.save(ATTENDANCE_FILE)
        return updated

    # ---------------- Session controls ----------------
    def _start_session(self):
        if self.active_session:
            messagebox.showinfo("Session already active", f"Session {self.active_session['session_id']} is already running.")
            return

        today_nodash = datetime.now().strftime("%Y%m%d")
        if self.session_state.get("date") != today_nodash:
            self.session_state = {"date": today_nodash, "next_session_number": 1, "active_session": None}

        n = self.session_state.get("next_session_number", 1)
        if n > 5:
            if not messagebox.askyesno(
                "More than 5 sessions today",
                f"This would be session {n} today (usual max is 5). Continue anyway?",
            ):
                return

        session_id = f"{today_nodash}_{n}"
        self.session_state["active_session"] = {"session_id": session_id, "start_time": datetime.now().isoformat()}
        save_session_state(self.session_state)
        self.active_session = self.session_state["active_session"]
        self._update_session_ui()

    def _end_session(self):
        if not self.active_session:
            messagebox.showinfo("No active session", "There is no session currently running.")
            return

        session_id = self.active_session["session_id"]
        result = ask_end_session_params(self, f"End Session {session_id}", self.last_instructor, "", self.last_hall)
        if result is None:
            return
        instructor, duration, hall = result
        self.last_instructor = instructor
        self.last_hall = hall

        updated_rows = self._apply_session_metadata(session_id, hall, instructor, duration)

        self.session_state["active_session"] = None
        self.session_state["next_session_number"] = self.session_state.get("next_session_number", 1) + 1
        save_session_state(self.session_state)
        self.active_session = None
        self._update_session_ui()

        messagebox.showinfo(
            "Session closed",
            f"Session {session_id} closed.\n"
            f"Hall: {hall}   TA: {instructor}   Duration: {duration:g} min\n"
            f"{updated_rows} scan row(s) tagged.\n\n"
            "Once every hall's file is merged, run attendance_analyzer.py to generate the report.",
        )

    def _update_session_ui(self):
        if self.active_session:
            self.session_status_label.config(text=f"Active session: {self.active_session['session_id']}")
            self.start_btn.config(state="disabled")
            self.end_btn.config(state="normal")
            self.att_entry.config(state="normal")
            self.att_entry.focus_set()
        else:
            self.session_status_label.config(text="No active session")
            self.start_btn.config(state="normal")
            self.end_btn.config(state="disabled")
            self.att_entry.config(state="disabled")

    # ---------------- Enroll tab ----------------
    def _build_enroll_tab(self):
        frame = self.enroll_tab
        ttk.Label(
            frame,
            text="1) Scan the card   2) Type the Student ID and click Search\n"
            "3) Found -> Name/Faculty auto-fill (still editable)   Not found -> enter them manually\n"
            "If a session is running, enrolling also logs this moment as that student's scan.",
            justify="left",
        ).pack(anchor="w", padx=8, pady=(8, 4))

        uid_row = ttk.Frame(frame)
        uid_row.pack(fill="x", padx=8, pady=4)
        ttk.Label(uid_row, text="Scan card:", width=14).pack(side="left")
        self.enroll_uid_var = tk.StringVar()
        self.enroll_uid_entry = ttk.Entry(uid_row, textvariable=self.enroll_uid_var, font=("Consolas", 12))
        self.enroll_uid_entry.pack(side="left", fill="x", expand=True)
        self.enroll_uid_entry.bind("<Return>", self._on_enroll_uid_scan)

        id_row = ttk.Frame(frame)
        id_row.pack(fill="x", padx=8, pady=4)
        ttk.Label(id_row, text="Student ID:", width=14).pack(side="left")
        self.enroll_id_var = tk.StringVar()
        self.enroll_id_entry = ttk.Entry(id_row, textvariable=self.enroll_id_var, font=("Consolas", 12))
        self.enroll_id_entry.pack(side="left", fill="x", expand=True)
        self.enroll_id_entry.bind("<Return>", lambda e: self._on_search_student())
        ttk.Button(id_row, text="Search", command=self._on_search_student).pack(side="left", padx=(6, 0))

        self.enroll_search_status = ttk.Label(frame, text="")
        self.enroll_search_status.pack(anchor="w", padx=8)

        name_row = ttk.Frame(frame)
        name_row.pack(fill="x", padx=8, pady=4)
        ttk.Label(name_row, text="Student name:", width=14).pack(side="left")
        self.enroll_name_var = tk.StringVar()
        self.enroll_name_entry = ttk.Entry(name_row, textvariable=self.enroll_name_var, font=("Consolas", 12))
        self.enroll_name_entry.pack(side="left", fill="x", expand=True)

        faculty_row = ttk.Frame(frame)
        faculty_row.pack(fill="x", padx=8, pady=4)
        ttk.Label(faculty_row, text="Faculty:", width=14).pack(side="left")
        self.enroll_faculty_var = tk.StringVar()
        self.enroll_faculty_combo = ttk.Combobox(
            faculty_row, textvariable=self.enroll_faculty_var, values=FACULTY_LIST, state="readonly", width=38
        )
        self.enroll_faculty_combo.pack(side="left", fill="x", expand=True)

        ttk.Button(frame, text="Enroll / Update Student", command=self._on_enroll_submit).pack(
            padx=8, pady=8, anchor="w"
        )

        self.enroll_list = tk.Listbox(frame, font=("Consolas", 11))
        self.enroll_list.pack(fill="both", expand=True, padx=8, pady=8)
        self._refresh_enroll_list()

        self.current_enroll_uid = None
        self.enroll_uid_entry.focus_set()

    def _on_enroll_uid_scan(self, event):
        raw = self.enroll_uid_var.get()
        uid = extract_uid(raw)
        if not uid:
            return
        self.current_enroll_uid = uid
        self.enroll_uid_var.set(uid)
        self.enroll_id_entry.focus_set()

    def _on_search_student(self):
        student_id = self.enroll_id_var.get().strip()
        if not student_id:
            messagebox.showwarning("Missing Student ID", "Type the Student ID first.")
            return

        match = self.student_db.get(student_id)
        if match:
            self.enroll_name_var.set(match["name"])
            self.enroll_faculty_var.set(match["faculty"])
            self.enroll_search_status.config(text="Found in the university database.", foreground="green")
        else:
            self.enroll_name_var.set("")
            self.enroll_faculty_var.set("")
            self.enroll_search_status.config(
                text="Not found in the database -- enter Name and Faculty manually.", foreground="red"
            )
        self.enroll_name_entry.focus_set()

    def _on_enroll_submit(self):
        uid = self.current_enroll_uid
        student_id = self.enroll_id_var.get().strip()
        name = self.enroll_name_var.get().strip()
        faculty = self.enroll_faculty_var.get().strip()

        if not uid:
            messagebox.showwarning("Missing card", "Scan a card in the 'Scan card' field first.")
            return
        if not student_id or not name or not faculty:
            messagebox.showwarning("Missing info", "Student ID, Name, and Faculty are all required.")
            return

        self.roster[uid] = {"student_id": student_id, "name": name, "faculty": faculty}
        save_roster(self.roster)
        self._refresh_enroll_list()

        if self.active_session:
            self._register_attendance(uid)
            note = f"Logged as attendance for session {self.active_session['session_id']}."
        else:
            note = "No active session right now, so attendance was not recorded."
        messagebox.showinfo("Student enrolled", f"{name} (ID {student_id}, {faculty}) enrolled.\n{note}")

        self.current_enroll_uid = None
        self.enroll_uid_var.set("")
        self.enroll_id_var.set("")
        self.enroll_name_var.set("")
        self.enroll_faculty_var.set("")
        self.enroll_search_status.config(text="")
        self.enroll_uid_entry.focus_set()

    def _refresh_enroll_list(self):
        self.enroll_list.delete(0, tk.END)
        for uid, info in self.roster.items():
            faculty = info.get("faculty", "")
            self.enroll_list.insert(
                tk.END, f"{uid}   ID: {info['student_id']:<10}   {info['name']:<25}   {faculty}"
            )

    # ---------------- Focus management ----------------
    # Focus is set explicitly right after actions that need it (a scan
    # processed, a session started, a tab switched) instead of a recurring
    # timer -- a polling timer was fighting with the Hall/Faculty dropdowns,
    # yanking focus back and closing them before a click could register.
    def _on_tab_changed(self, event=None):
        try:
            current_tab = self.nametowidget(self.notebook.select())
            if current_tab == self.attendance_tab and str(self.att_entry.cget("state")) == "normal":
                self.att_entry.focus_set()
            elif current_tab == self.enroll_tab:
                self.enroll_uid_entry.focus_set()
        except Exception:
            pass


if __name__ == "__main__":
    App().mainloop()
